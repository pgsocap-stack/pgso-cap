import tkinter as tk
from tkinter import ttk, messagebox
import db  # Ginagamit para humila ng data sa database layer
import os  # 👈 Para sa paggawa ng folders
from datetime import datetime  # 👈 Para sa real-time date

def open_excel_preview_modal(parent_window, project_treeview):
    """
    Ito ang Main Entry Point ng Module.
    Nagluluwal ng isang Popup Window para sa Print Preview at Excel Saving.
    """
    
    # ==========================================================================
    # DETALYE 1: PAGKUHA NG DETALYE NG NAPILING PROYEKTO (UPDATED CODES)
    # ==========================================================================
    try:
        selected_project = project_treeview.selection() 
        if not selected_project:
            messagebox.showwarning("Walang Napili", "Pumili muna ng proyekto sa listahan bago mag-preview.")
            return
        
        row_values = project_treeview.item(selected_project[0], 'values')
        pow_id = row_values[2] 
        
    except Exception as e:
        messagebox.showwarning("Selection Error", f"Hindi makuha ang detalye ng napiling proyekto: {e}")
        return

    proj_info = db.get_project_details(pow_id)
    pow_items = db.get_items_by_project(pow_id)

    if not proj_info:
        messagebox.showerror("Error", "Hindi mahanap ang detalye ng proyektong ito sa database.")
        return

    project_name, location = proj_info[0], proj_info[1]

    # ==========================================================================
    # DETALYE 2: PAGDISENYO AT PAGSUKAT NG PREVIEW POPUP WINDOW
    # ==========================================================================
    preview_win = tk.Toplevel(parent_window)
    preview_win.title("POW - Print & Office Excel Layout Preview")
    
    screen_width = preview_win.winfo_screenwidth()
    screen_height = preview_win.winfo_screenheight()
    
    win_width = int(screen_width / 2)       
    win_height = int(screen_height * 0.85)   
    
    start_x = int((screen_width - win_width) / 2)
    start_y = int((screen_height - win_height) / 2)
    preview_win.geometry(f"{win_width}x{win_height}+{start_x}+{start_y}")
    
    preview_win.transient(parent_window.winfo_toplevel())
    preview_win.grab_set()

    top_bar = tk.Frame(preview_win, bg="#edf2f7", padx=15, pady=10)
    top_bar.pack(fill="x", side="top")

    text_frame = tk.Frame(preview_win, bg="white", padx=20, pady=20)
    text_frame.pack(fill="both", expand=True)

    scroll_y = ttk.Scrollbar(text_frame, orient="vertical")
    scroll_y.pack(fill="y", side="right")
    
    scroll_x = ttk.Scrollbar(text_frame, orient="horizontal")
    scroll_x.pack(fill="x", side="bottom")

    preview_text = tk.Text(text_frame, font=("Consolas", 10), bg="#ffffff", fg="#2d3748", 
                            wrap="none", yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
    preview_text.pack(fill="both", expand=True, side="left")
    
    scroll_y.config(command=preview_text.yview)
    scroll_x.config(command=preview_text.xview)

    # ==========================================================================
    # DETALYE 3: PAGBUO NG SAKTONG ANYO NG PRINT PREVIEW (TEXT FORMAT)
    # ==========================================================================
    lines = []
    lines.append(f"{'':<25}Republic of the Philippines")
    lines.append(f"{'':<22}PROVINCE   OF   NUEVA   ECIJA")
    lines.append(f"{'':<26}Palayan City")
    lines.append(f"{'':<15}PROVINCIAL GENERAL SERVICES OFFICE")
    lines.append("")
    lines.append(f"{'':<27}PROGRAM OF WORKS")
    lines.append("")
    lines.append(f"Project:  {project_name}")
    lines.append(f"Location: {location}")
    lines.append("")
    
    lines.append("=" * 95)
    lines.append(f"{'ITEM':<8}{'QTY':<8}{'UNIT':<10}{'DESCRIPTION':<35}{'UNIT PRICE':<16}{'AMOUNT':<15}")
    lines.append("=" * 95)

    grand_total = 0.0
    for idx, item in enumerate(pow_items, start=1):
        qty = float(item[0]) # 👈 GINAWANG FLOAT PARA SA RE-TEXT MULTIPLICATION!
        unit = item[1]
        raw_name = item[2]
        raw_price = item[3]
        
        name = str(raw_name).replace("\n", " ").replace("\r", " ").strip()
        try:
            price = float(raw_price) if raw_price is not None else 0.0
        except ValueError:
            price = 0.0

        amount = qty * price
        grand_total += amount
        
        short_name = name[:32] + "..." if len(name) > 32 else name
        lines.append(f"{idx:<8}{qty:<8.2f}{unit:<10}{short_name:<35}{price:>12,.2f}      {amount:>12,.2f}")

    lines.append("-" * 95)
    lines.append(f"{'TOTAL':<61}P     {grand_total:>22,.2f}")
    lines.append("=" * 95)
    lines.append("")

    lines.append(f"Prepared by:{'':<45}Checked by:")
    lines.append("")
    lines.append(f"       JONATHAN G. LADIGNON{'':<37}BENJAMIN N. RAMOS JR")
    lines.append(f"       Admin. Officer III  {'':<37}Engineer II")
    lines.append("")
    lines.append(f"Noted by:{'':<48}Recommending Approval:")
    lines.append("")
    lines.append(f"MARIO T. MARIANO{'':<41}ENGR. FLORECIO M. VALINO")
    lines.append("")
    lines.append(f"{'':<50}Approved:")
    lines.append("")
    lines.append(f"{'':<45}HON. AURELIO M. UMALI")
    lines.append(f"{'':<50}Governor")

    for line in lines:
        preview_text.insert(tk.END, line + "\n")
    preview_text.config(state="disabled")

    # ==========================================================================
    # DETALYE 4: CODES PARA SA AUTOMATED EXCEL FILE SAVING (openpyxl ENGINE)
    # ==========================================================================
    def export_to_actual_excel():
        """Lohika na gumagawa at nag-o-automate ng totoong Microsoft Excel file."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Program of Work"
        ws.views.sheetView[0].showGridLines = True

        # 📑 PRINT & PAGE SETUP FOR LEGAL SIZE
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.paperSize = ws.PAPERSIZE_LEGAL
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1   
        ws.page_setup.fitToHeight = 0  

        ws.page_margins.left = 0.5
        ws.page_margins.right = 0.5
        ws.page_margins.top = 0.75
        ws.page_margins.bottom = 0.75

        font_title = Font(name="Arial", size=10, bold=True)
        font_regular = Font(name="Arial", size=10)
        font_bold_body = Font(name="Arial", size=10, bold=True)
        
        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")

        ws['A1'] = "Republic of the Philippines"
        ws['A2'] = "PROVINCE   OF   NUEVA   ECIJA"
        ws['A3'] = "Palayan City"
        ws['A4'] = "PROVINCIAL GENERAL SERVICES OFFICE"
        ws['A6'] = "PROGRAM OF WORKS"
        
        for r in range(1, 7):
            if ws.cell(row=r, column=1).value:
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
                header_cell = ws.cell(row=r, column=1)
                header_cell.alignment = align_center
                header_cell.font = font_title

        ws['A7'] = f"Project: {project_name}"
        ws['A8'] = f"Location: {location}"
        ws['A7'].font = font_bold_body
        ws['A8'].font = font_bold_body

        headers_list = ["ITEM", "QTY", "UNIT", "DESCRIPTION", "UNIT PRICE", "AMOUNT"]
        for col_num, header_text in enumerate(headers_list, start=1):
            c_cell = ws.cell(row=9, column=col_num, value=header_text)
            c_cell.font = font_bold_body
            c_cell.alignment = align_center
            
        row_tracker = 10
        computed_total = 0.0
        
        for idx, item in enumerate(pow_items, start=1):
            qty = float(item[0]) # 👈 GINAWANG FLOAT DIN DITO PARA SA EXCEL SAVER ENGINE!
            unit = item[1]
            raw_name = item[2]
            raw_price = item[3]
            
            name = str(raw_name).replace("\n", " ").replace("\r", " ").strip()
            try:
                price = float(raw_price) if raw_price is not None else 0.0
            except ValueError:
                price = 0.0

            amount = qty * price
            computed_total += amount

            ws.cell(row=row_tracker, column=1, value=idx).alignment = align_center
            ws.cell(row=row_tracker, column=2, value=qty).alignment = align_center
            ws.cell(row=row_tracker, column=3, value=unit).alignment = align_center
            ws.cell(row=row_tracker, column=4, value=name).alignment = align_left
            
            p_cell = ws.cell(row=row_tracker, column=5, value=price)
            p_cell.number_format = '#,##0.00'
            p_cell.alignment = align_right
            
            a_cell = ws.cell(row=row_tracker, column=6, value=amount)
            a_cell.number_format = '#,##0.00'
            a_cell.alignment = align_right

            for column_pos in range(1, 7):
                ws.cell(row=row_tracker, column=column_pos).font = font_regular
                
            row_tracker += 1

        row_tracker += 1 
        ws.cell(row=row_tracker, column=5, value="Total  P").font = font_bold_body
        ws.cell(row=row_tracker, column=5).alignment = Alignment(horizontal="right", vertical="center")
        
        f_total_cell = ws.cell(row=row_tracker, column=6, value=computed_total)
        f_total_cell.font = font_bold_body
        f_total_cell.number_format = '"P" #,##0.00'
        f_total_cell.alignment = align_right

        double_bottom_border = Border(bottom=Side(style='double'))
        last_item_amount_cell = ws.cell(row=row_tracker - 2, column=6)
        last_item_amount_cell.border = double_bottom_border

        row_tracker += 2
        ws.cell(row=row_tracker, column=1, value="Prepared by:                                                                                Checked by:").font = font_regular
        
        row_tracker += 2
        ws.cell(row=row_tracker, column=1, value="        JONATHAN G. LADIGNON                                                                BENJAMIN N. RAMOS JR.").font = font_bold_body
        
        row_tracker += 1
        ws.cell(row=row_tracker, column=1, value="             Admin. Officer III                                                                                  Engineer II").font = font_regular

        row_tracker += 2
        ws.cell(row=row_tracker, column=1, value="Noted by:                                                                                     Recommending Approval:").font = font_regular

        row_tracker += 2
        ws.cell(row=row_tracker, column=1, value="        MARIO T. MARIANO                                                                     ENGR. FLORECIO M. VALINO").font = font_bold_body
        
        row_tracker += 1
        ws.cell(row=row_tracker, column=1, value="             Engineer IV                                                                                          PGS-Officer").font = font_regular

        row_tracker += 2
        ws.cell(row=row_tracker, column=4, value="                             Approved:").font = font_regular

        row_tracker += 2
        ws.cell(row=row_tracker, column=4, value="                     HON. AURELIO M. UMALI").font = font_bold_body
        row_tracker += 1
        ws.cell(row=row_tracker, column=4, value="                                 Governor").font = font_regular

        column_widths = {'A': 4.57, 'B': 4.86, 'C': 7.00, 'D': 47.14, 'E': 11.57, 'F': 14.57}
        for col_letter, width_size in column_widths.items():
            ws.column_dimensions[col_letter].width = width_size

        ws.row_dimensions[9].height = 20
        for r_idx in range(10, row_tracker + 15):
            ws.row_dimensions[r_idx].height = 16

        if len(pow_items) > 35:
            from openpyxl.worksheet.pagebreak import Break
            ws.row_breaks.append(Break(id=45))

        # ==========================================================================
        # 📂 AUTOMATED USB FOLDER & SAVE CONFIGURATION
        # ==========================================================================
        date_folder_name = datetime.now().strftime("%m-%Y")
        usb_base_path = r"G:\jrm"
        
        if os.path.exists("G:\\"):
            final_export_dir = os.path.join(usb_base_path, date_folder_name)
            os.makedirs(final_export_dir, exist_ok=True)
            usb_success = True
        else:
            final_export_dir = os.getcwd()
            usb_success = False

        clean_project_name = project_name.replace(' ', '_').replace('/', '-').replace('\\', '-')
        filename = f"POW_{clean_project_name}.xlsx"
        full_save_path = os.path.join(final_export_dir, filename)

        try:
            wb.save(full_save_path)
            if usb_success:
                messagebox.showinfo("Tagumpay", f"Matagumpay na nai-save sa USB, boss!\n\nNaka-sort sa Monthly Folder:\n{full_save_path}")
            else:
                messagebox.showwarning("Walang USB Drive", f"Babala: Hindi nahanap ang G:\\ drive (USB).\n\nPansamantala itong isinave sa System Folder:\n\n{filename}")
        except Exception as ex:
            messagebox.showerror("Error sa Pag-save", f"Hindi mai-save ang file. Siguraduhing sarado ang lumang Excel file.\nDetalye: {ex}")

    # Pagkakabit ng Control Buttons sa Top Action Bar ng popup UI
    btn_save_excel = tk.Button(top_bar, text="📥 EXPORT to EXCEL FILE", bg="#28a745", fg="white", 
                               font=("Segoe UI", 10, "bold"), padx=15, command=export_to_actual_excel, cursor="hand2")
    btn_save_excel.pack(side="left")
    
    btn_quit_preview = tk.Button(top_bar, text="❌ CANCEL", bg="#dc3545", fg="white", 
                                 font=("Segoe UI", 10, "bold"), padx=15, command=preview_win.destroy, cursor="hand2")
    btn_quit_preview.pack(side="right")