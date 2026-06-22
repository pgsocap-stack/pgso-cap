import subprocess
from openpyxl import load_workbook
import excel_generator  # 👈 Idagdag ito sa pinaka-itaas ng UI file mo
import tkinter as tk
from tkinter import ttk, messagebox
import db

class PreviewPowModule(tk.Frame):
    def __init__(self, parent):
        super().__init__(parent, bg="#f7fafc")
        
        self.current_selected_pow_id = None # Para malaman kung anong idedelete

        # --- HEADER ---
        header = tk.Label(self, text="PREVIEW SAVED PROGRAM OF WORK (POW)", font=("Segoe UI", 16, "bold"), fg="#1a365d", bg="#f7fafc")
        header.pack(anchor="w", padx=20, pady=15)

        # --- MAIN SPLIT CONTAINER ---
        main_pane = tk.PanedWindow(self, orient="horizontal", bg="#e2e8f0")
        main_pane.pack(fill="both", expand=True, padx=20, pady=10)

        # 👈 KALIWA: Project Selection
        left_frame = tk.LabelFrame(main_pane, text=" List of Projects ", font=("Segoe UI", 10, "bold"), bg="white", width=300)
        left_frame.pack_propagate(False)
        main_pane.add(left_frame)

        self.proj_tree = ttk.Treeview(left_frame, columns=("ID", "Project Name"), show="headings")
        self.proj_tree.heading("ID", text="ID")
        self.proj_tree.heading("Project Name", text="Project Title")
        self.proj_tree.column("ID", width=40, anchor="center")
        self.proj_tree.column("Project Name", width=240, anchor="w")
        self.proj_tree.pack(fill="both", expand=True, padx=5, pady=5)
        self.proj_tree.bind("<<TreeviewSelect>>", self.on_project_select)

        # 🗑️ BUTTON SA KALIWA: Delete Entire POW
        self.btn_delete_pow = tk.Button(left_frame, text="❌ Delete Selected POW", bg="#e53e3e", fg="white", font=("Segoe UI", 9, "bold"), pady=5, command=self.delete_entire_project, state="disabled", cursor="hand2")
        self.btn_delete_pow.pack(fill="x", padx=5, pady=5)

        # 👉 KANAN: Items List
        right_frame = tk.LabelFrame(main_pane, text=" Item list on POW ", font=("Segoe UI", 10, "bold"), bg="white")
        main_pane.add(right_frame)

        self.lbl_location = tk.Label(right_frame, text="Location: (Select Project)", font=("Segoe UI", 11, "italic"), fg="#4a5568", bg="white", anchor="w")
        self.lbl_location.pack(fill="x", padx=10, pady=10)

        self.items_tree = ttk.Treeview(right_frame, columns=("No", "Qty", "Unit", "Item Description", "Price", "Total"), show="headings")
        self.items_tree.heading("No", text="#")
        self.items_tree.heading("Qty", text="Qty")
        self.items_tree.heading("Unit", text="Unit")
        self.items_tree.heading("Item Description", text="Item Description")
        self.items_tree.heading("Price", text="Unit Price")
        self.items_tree.heading("Total", text="Total Price")
        
        self.items_tree.column("No", width=35, anchor="center")
        self.items_tree.column("Qty", width=50, anchor="center")
        self.items_tree.column("Unit", width=60, anchor="center")
        self.items_tree.column("Item Description", width=250, anchor="w")
        self.items_tree.column("Price", width=90, anchor="e")
        self.items_tree.column("Total", width=100, anchor="e")
        self.items_tree.pack(fill="both", expand=True, padx=5, pady=5)
        
        # 🟢 CONTAINER PARA SA MGA BUTTONS SA ILALIM NG TABLE
        bottom_button_frame = tk.Frame(right_frame, bg="white", pady=10)
        bottom_button_frame.pack(fill="x", side="bottom")
        
        # 👁️ ANG PREVIEW DATA BUTTON
        self.preview_btn = tk.Button(
            bottom_button_frame, 
            text="👁️ Preview Data", 
            bg="#3182ce", 
            fg="white", 
            font=("Segoe UI", 10, "bold"), 
            padx=15,
            command=lambda: excel_generator.open_excel_preview_modal(self, self.proj_tree)
        )
        self.preview_btn.pack(side="left", padx=10)
        
        # 🔴 ANG BAGONG PULANG EDIT BUTTON
        self.btn_edit_pow = tk.Button(
            bottom_button_frame,   
            text="✏️ Edit POW Record",
            font=("Segoe UI", 10, "bold"),
            bg="#e53e3e",          
            fg="white",            
            activebackground="#c53030",
            activeforeground="white",
            state="disabled",      
            cursor="hand2",
            padx=15,
            command=self.open_edit_pow_modal 
        )
        self.btn_edit_pow.pack(side="right", padx=10)

        # 💰 SUMMARY BAR SA KANAN
        preview_summary = tk.Frame(right_frame, bg="#e2e8f0", padx=15, pady=8)
        preview_summary.pack(fill="x", padx=5, pady=5)
        
        self.lbl_preview_total = tk.Label(preview_summary, text="PROJECT TOTAL COST: P 0.00", font=("Segoe UI", 11, "bold"), fg="#2c5282", bg="#e2e8f0")
        self.lbl_preview_total.pack(side="right")

        self.load_projects_from_db()

    def load_projects_from_db(self):
        for row in self.proj_tree.get_children():
            self.proj_tree.delete(row)
            
        projects = db.get_project_list()
        for index, proj in enumerate(projects, start=1):
            pow_id = proj[0]        
            project_name = proj[1]  
            location = proj[2]      
            self.proj_tree.insert("", "end", values=(index, project_name, pow_id), tags=(location,))
            
        self.btn_delete_pow.config(state="disabled") 
        self.btn_edit_pow.config(state="disabled")   
        self.lbl_preview_total.config(text="PROJECT TOTAL COST: P 0.00")

    def on_project_select(self, event):
        selected_item = self.proj_tree.selection()
        if not selected_item: return

        row_values = self.proj_tree.item(selected_item[0], 'values')
        self.current_selected_pow_id = row_values[2]
        location = self.proj_tree.item(selected_item[0], 'tags')[0]
        
        self.lbl_location.config(text=f"📍 Location: {location}", font=("Segoe UI", 11, "bold"), fg="#2b6cb0")
        self.btn_delete_pow.config(state="normal") 
        self.btn_edit_pow.config(state="normal")   

        for row in self.items_tree.get_children():
            self.items_tree.delete(row)

        associated_items = db.get_items_by_project(self.current_selected_pow_id)
        grand_total = 0.0

        for index, item in enumerate(associated_items, start=1):
            qty = float(item[0]) 
            unit = item[1]
            name = item[2]
            price = float(item[3])
            total = qty * price
            grand_total += total
            self.items_tree.insert("", "end", values=(index, qty, unit, name, f"P {price:,.2f}", f"P {total:,.2f}"))
            
        self.lbl_preview_total.config(text=f"PROJECT TOTAL COST: P {grand_total:,.2f}")

    def delete_entire_project(self):
        if not self.current_selected_pow_id: return
        confirm = messagebox.askyesno("Confirm Delete", "Sigurado ka bang buburahin ang buong proyektong ito?\n\nHindi na ito mababawi.")
        if confirm:
            success = db.delete_pow_from_sql(self.current_selected_pow_id)
            if success:
                messagebox.showinfo("Deleted", "Matagumpay na nabura ang buong POW record.")
                for row in self.items_tree.get_children(): self.items_tree.delete(row)
                self.lbl_location.config(text="Location: (Select Project)", font=("Segoe UI", 11, "italic"), fg="#4a5568")
                self.load_projects_from_db()

    def open_edit_pow_modal(self):
        if not self.current_selected_pow_id: return

        selected_item = self.proj_tree.selection()[0]
        row_values = self.proj_tree.item(selected_item, 'values')
        
        current_project_name = row_values[1] 
        current_location = self.proj_tree.item(selected_item, 'tags')[0] 
        associated_items = db.get_items_by_project(self.current_selected_pow_id)

        edit_win = tk.Toplevel(self)
        edit_win.title("✏️ EDIT MODE - Update POW Record")
        edit_win.geometry("950x650")
        edit_win.configure(bg="#f7fafc")
        edit_win.grab_set() 

        # --- UPPER FRAME: DETAILS ---
        top_frame = tk.LabelFrame(edit_win, text=" Details of POW ", font=("Segoe UI", 10, "bold"), bg="#f7fafc", fg="#2d3748", padx=15, pady=10)
        top_frame.pack(fill="x", padx=15, pady=10)

        tk.Label(top_frame, text="Project Title / Name:", font=("Segoe UI", 10), bg="#f7fafc").grid(row=0, column=0, sticky="w", pady=5)
        ent_proj_name = tk.Entry(top_frame, font=("Segoe UI", 11), width=50)
        ent_proj_name.grid(row=0, column=1, padx=10, pady=5, sticky="w")
        ent_proj_name.insert(0, current_project_name) 

        tk.Label(top_frame, text="Project Location:", font=("Segoe UI", 10), bg="#f7fafc").grid(row=1, column=0, sticky="w", pady=5)
        ent_location = tk.Entry(top_frame, font=("Segoe UI", 11), width=50)
        ent_location.grid(row=1, column=1, padx=10, pady=5, sticky="w")
        ent_location.insert(0, current_location) 

        # --- MIDDLE FRAME: TABLE ---
        items_frame = tk.LabelFrame(edit_win, text=" List of Items ", font=("Segoe UI", 10, "bold"), bg="#f7fafc", fg="#2d3748", padx=15, pady=10)
        items_frame.pack(fill="both", expand=True, padx=15, pady=5)

        # 💡 DAGDAG KOLUM: Nilagyan natin ng 'orig_desc' para matandaan ang lumang pangalan kahit i-edit mo pa!
        columns = ("qty", "unit", "description", "price", "orig_desc")
        edit_tree = ttk.Treeview(items_frame, columns=columns, show="headings", height=10)
        edit_tree.heading("qty", text="QTY")
        edit_tree.heading("unit", text="UNIT")
        edit_tree.heading("description", text="ITEM DESCRIPTION")
        edit_tree.heading("price", text="UNIT PRICE")
        edit_tree.heading("orig_desc", text="ORIGINAL NAME") # Hidden sa mata ng user mamaya
        
        edit_tree.column("qty", width=70, anchor="center")
        edit_tree.column("unit", width=70, anchor="center")
        edit_tree.column("description", width=400, anchor="w")
        edit_tree.column("price", width=120, anchor="e")
        edit_tree.column("orig_desc", width=0, stretch=tk.NO) # 👈 Tinago natin para hindi makalat sa UI
        edit_tree.pack(fill="both", expand=True, side="left")

        sb = ttk.Scrollbar(items_frame, orient="vertical", command=edit_tree.yview)
        sb.pack(fill="y", side="right")
        edit_tree.configure(yscrollcommand=sb.set)

        for item in associated_items:
            # Pagkapasok, parehong item description ang nilalagay sa description at orig_desc column
            edit_tree.insert("", "end", values=(item[0], item[1], item[2], f"{float(item[3]):.2f}", item[2]))

        # --- FORM INPUTS FOR ROWS ---
        form_frame = tk.Frame(edit_win, bg="#f7fafc")
        form_frame.pack(fill="x", padx=15, pady=10)

        tk.Label(form_frame, text="QTY:", bg="#f7fafc").grid(row=0, column=0, padx=2)
        ent_qty = tk.Entry(form_frame, width=8, font=("Segoe UI", 10))
        ent_qty.grid(row=0, column=1, padx=5)

        tk.Label(form_frame, text="UNIT:", bg="#f7fafc").grid(row=0, column=2, padx=2)
        ent_unit = tk.Entry(form_frame, width=8, font=("Segoe UI", 10))
        ent_unit.grid(row=0, column=3, padx=5)

        tk.Label(form_frame, text="DESCRIPTION:", bg="#f7fafc").grid(row=0, column=4, padx=2)
        ent_desc = tk.Entry(form_frame, width=35, font=("Segoe UI", 10))
        ent_desc.grid(row=0, column=5, padx=5)

        tk.Label(form_frame, text="PRICE:", bg="#f7fafc").grid(row=0, column=6, padx=2)
        ent_price = tk.Entry(form_frame, width=12, font=("Segoe UI", 10))
        ent_price.grid(row=0, column=7, padx=5)

        # Gagawa tayo ng variable para matandaan kung anong orig name ang kasalukuyang ine-edit sa entry boxes
        self.current_editing_orig_desc = ""

        def on_edit_row_select(event):
            sel = edit_tree.selection()
            if not sel: return
            v = edit_tree.item(sel[0], 'values')
            ent_qty.delete(0, tk.END); ent_qty.insert(0, v[0])
            ent_unit.delete(0, tk.END); ent_unit.insert(0, v[1])
            ent_desc.delete(0, tk.END); ent_desc.insert(0, v[2])
            ent_price.delete(0, tk.END); ent_price.insert(0, v[3])
            # Kunin ang natagong Original Name mula sa ika-5 kolum (index 4)
            self.current_editing_orig_desc = v[4]

        edit_tree.bind("<<TreeviewSelect>>", on_edit_row_select)

        def apply_row_change():
            sel = edit_tree.selection()
            if not sel:
                messagebox.showwarning("Puna", "Pumili muna ng linya sa table, boss.")
                return
            # Pag in-update ang linya, pinapanatili pa rin natin ang original name nito sa dulo para sa Excel later
            edit_tree.item(sel[0], values=(ent_qty.get(), ent_unit.get(), ent_desc.get(), ent_price.get(), self.current_editing_orig_desc))
            clear_entry_fields()

        def add_new_row():
            if not ent_desc.get():
                messagebox.showwarning("Puna", "Lagyan ng Description ang bagong aytem, boss.")
                return
            qty_val = ent_qty.get() if ent_qty.get() else "0"
            price_val = ent_price.get() if ent_price.get() else "0.00"
            # Kung talagang sariwang bagong item ito, ang orig_desc nito ay kapareho rin ng bagong description
            edit_tree.insert("", "end", values=(qty_val, ent_unit.get(), ent_desc.get(), price_val, ent_desc.get()))
            clear_entry_fields()

        def clear_entry_fields():
            ent_qty.delete(0, tk.END); ent_unit.delete(0, tk.END)
            ent_desc.delete(0, tk.END); ent_price.delete(0, tk.END)
            self.current_editing_orig_desc = ""

        btn_apply = tk.Button(form_frame, text="🔄 Update Selected Line", bg="#3182ce", fg="white", font=("Segoe UI", 9, "bold"), command=apply_row_change)
        btn_apply.grid(row=0, column=8, padx=3)

        btn_add_line = tk.Button(form_frame, text="➕ Add Line", bg="#38a169", fg="white", font=("Segoe UI", 9, "bold"), command=add_new_row)
        btn_add_line.grid(row=0, column=9, padx=3)

        # --- FINAL SAVE OVERWRITE PIPELINE ---
        bottom_frame = tk.Frame(edit_win, bg="#f7fafc")
        bottom_frame.pack(fill="x", side="bottom", pady=15)

        def save_all_updates_to_db():
            new_name = ent_proj_name.get().strip()
            new_loc = ent_location.get().strip()
            
            if not new_name or not new_loc:
                messagebox.showerror("Error", "Huwag iwanang blangko ang Name at Location, boss.")
                return

            final_items_to_save = []
            excel_path = r"G:\jrm\master_items.xlsx"

            confirm = messagebox.askyesno("Kumpirmasyon", "Gusto mo bang ituloy ang pag-save gamit ang Excel-Database Sync pipeline?")
            if not confirm: return

            try:
                wb = load_workbook(excel_path)
                ws = wb.active
            except Exception as e:
                messagebox.showerror("Excel Open Error", f"Hindi mabuksan ang {excel_path}.\nDetalye: {e}")
                return

            # LOOP sa bawat item na nasa edit_tree gamit ang matalinong TRACER LOGIC
            for row in edit_tree.get_children():
                vals = edit_tree.item(row, 'values')
                try:
                    q = float(vals[0])
                    u = str(vals[1])
                    d = str(vals[2]).strip()
                    p = float(vals[3])
                    orig_d = str(vals[4]).strip() # 🎯 Heto ang Lumang Pangalan!
                    
                    final_items_to_save.append((q, u, d, p))

                    # 🔍 TRACER HAKBANG: Sa Excel, ang hahanapin natin ay ang LUNANG PANGALAN (orig_d)
                    target_row = None
                    for ex_row in range(2, ws.max_row + 1):
                        cell_val = ws.cell(row=ex_row, column=1).value 
                        if cell_val and str(cell_val).strip().lower() == orig_d.lower():
                            target_row = ex_row
                            break

                    if target_row:
                        # 💥 OVERWRITE SUCCESS: Palitan ang buong row ng mga BAGONG detalye pati ang bagong Description!
                        ws.cell(row=target_row, column=1, value=d) # Bagong Description (May wood na)
                        ws.cell(row=target_row, column=2, value=u) # Bagong Unit
                        ws.cell(row=target_row, column=3, value=p) # Bagong Presyo
                    else:
                        # Kung totoong bagong item (Add Line), ilagay sa pinakailalim
                        new_row = ws.max_row + 1
                        ws.cell(row=new_row, column=1, value=d)
                        ws.cell(row=new_row, column=2, value=u)
                        ws.cell(row=new_row, column=3, value=p)

                except ValueError:
                    messagebox.showerror("Format Error", "Dapat numero ang QTY at PRICE, boss.")
                    wb.close()
                    return

            try:
                wb.save(excel_path)
                wb.close()
            except Exception as e:
                messagebox.showerror("Excel Save Error", f"Hindi mai-save ang mga bagong item sa Excel.\nDetalye: {e}")
                return

            # Background Automation Import
            try:
                subprocess.run(["py", "import_master.py"], cwd=r"G:\jrm", check=True, capture_output=True, text=True)
            except Exception as err:
                messagebox.showerror("Import Script Error", f"Na-save sa Excel pero nabigong patakbuhin ang 'import_master.py'.\n\nDetalye: {err}")
                return

            # Pag-update sa local POW server
            success_main = db.update_project_main_details(self.current_selected_pow_id, new_name, new_loc)
            success_items = db.update_project_items_batch(self.current_selected_pow_id, final_items_to_save)

            if success_main and success_items:
                messagebox.showinfo("Tagumpay", "Swabe ang ikot, boss!\n\n1. Napalitan ang lumang Record sa Excel\n2. Na-update ang Master List sa SQL\n3. Refresh Complete!")
                edit_win.destroy() 
                self.load_projects_from_db() 
                for r in self.items_tree.get_children(): self.items_tree.delete(r)
                self.lbl_location.config(text="Location: (Pumili muna ng proyekto)", font=("Segoe UI", 11, "italic"), fg="#4a5568")
            else:
                messagebox.showerror("SQL Error", "May error sa pag-update ng POW tables sa database.")

        btn_final_save = tk.Button(bottom_frame, text="💾 SAVE ALL CHANGES & OVERWRITE DATABASE", font=("Segoe UI", 11, "bold"), bg="#e53e3e", fg="white", padx=20, pady=8, command=save_all_updates_to_db)
        btn_final_save.pack(anchor="center")