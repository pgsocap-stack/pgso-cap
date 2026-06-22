import subprocess
from openpyxl import load_workbook
import tkinter as tk
from tkinter import ttk, messagebox
import db

class AddPowModule(tk.Frame):
    def __init__(self, parent):
        super().__init__(parent, bg="#f7fafc")
        
        self.temporary_items = []
        self.master_items_cache = {} # Itatago dito ang Unit at Price ng mga dropdown choices
        self.search_popup = None     # Dito itatago ang floating window ng search results

        # --- HEADER ---
        header = tk.Label(self, text="CREATE PROGRAM OF WORK (POW)", font=("Segoe UI", 16, "bold"), fg="#1a365d", bg="#f7fafc")
        header.pack(anchor="w", padx=20, pady=15)

        # --- FORM INPUTS ---
        form_frame = tk.LabelFrame(self, text=" Item Entry (Max 150 Items) ", font=("Segoe UI", 10, "bold"), bg="white", padx=15, pady=15)
        form_frame.pack(fill="x", padx=20, pady=10)

        tk.Label(form_frame, text="Qty:", bg="white").grid(row=0, column=0, sticky="e", pady=5)
        self.qty_entry = tk.Entry(form_frame, width=10)
        self.qty_entry.grid(row=0, column=1, padx=5, sticky="w")
        self.qty_entry.insert(0, "0") # Default value para iwas ValueError

        tk.Label(form_frame, text="Unit:", bg="white").grid(row=0, column=2, sticky="e", pady=5)
        self.unit_combobox = ttk.Combobox(form_frame, values=["gal", "lit", "tin", "pc", "box", "bag", "cu.m"], width=10)
        self.unit_combobox.grid(row=0, column=3, padx=5, sticky="w")
        self.unit_combobox.set("pc")

        tk.Label(form_frame, text="Item Name:", bg="white").grid(row=0, column=4, sticky="e", pady=5)
        
        self.name_entry = tk.Entry(form_frame, width=35)
        self.name_entry.grid(row=0, column=5, padx=5, sticky="w")
        
        self.name_entry.bind("<Return>", self.trigger_multi_dynamic_search) # Enter sa Item Name field

        tk.Label(form_frame, text="Unit Price:", bg="white").grid(row=0, column=6, sticky="e", pady=5)
        self.price_entry = tk.Entry(form_frame, width=12)
        self.price_entry.grid(row=0, column=7, padx=5, sticky="w")
        self.price_entry.insert(0, "0.00") # Default value para iwas ValueError

        self.btn_add_item = tk.Button(form_frame, text="➕ Add to List", bg="#3182ce", fg="white", font=("Segoe UI", 9, "bold"), command=self.add_item_to_list, cursor="hand2")
        self.btn_add_item.grid(row=0, column=8, padx=15)

        self.price_entry.bind("<Return>", lambda event: self.add_item_to_list())
        self.qty_entry.bind("<Return>", lambda event: self.add_item_to_list()) # Enter sa Qty field

        # --- TABLE PREVIEW ---
        table_frame = tk.Frame(self, bg="white")
        table_frame.pack(fill="both", expand=True, padx=20, pady=10)

        self.tree = ttk.Treeview(table_frame, columns=("No", "Qty", "Unit", "Item Name", "Price", "Total"), show="headings")
        self.tree.heading("No", text="#")
        self.tree.heading("Qty", text="Qty")
        self.tree.heading("Unit", text="Unit")
        self.tree.heading("Item Name", text="Item Name")
        self.tree.heading("Price", text="Unit Price")
        self.tree.heading("Total", text="Total Price")
        
        self.tree.column("No", width=40, anchor="center")
        self.tree.column("Qty", width=60, anchor="center")
        self.tree.column("Unit", width=80, anchor="center")
        self.tree.column("Item Name", width=350, anchor="w")
        self.tree.column("Price", width=100, anchor="e")
        self.tree.column("Total", width=120, anchor="e")
        self.tree.pack(fill="both", expand=True, side="left")

        scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(fill="y", side="right")

        # --- SUMMARY PANEL ---
        summary_frame = tk.Frame(self, bg="#edf2f7", padx=15, pady=10)
        summary_frame.pack(fill="x", padx=20, pady=5)

        self.lbl_grand_total = tk.Label(summary_frame, text="GRAND TOTAL: P 0.00", font=("Segoe UI", 12, "bold"), fg="#2b6cb0", bg="#edf2f7")
        self.lbl_grand_total.pack(side="right")

        btn_remove_item = tk.Button(summary_frame, text="🗑️ Remove Selected Item", bg="#e53e3e", fg="white", font=("Segoe UI", 9, "bold"), command=self.remove_selected_item, cursor="hand2")
        btn_remove_item.pack(side="left")

        # --- ACTION BUTTONS ---
        actions_frame = tk.Frame(self, bg="#f7fafc")
        actions_frame.pack(fill="x", padx=20, pady=10)

        self.lbl_counter = tk.Label(actions_frame, text="Total Items: 0 / 150", font=("Segoe UI", 10, "bold"), fg="#4a5568", bg="#f7fafc")
        self.lbl_counter.pack(side="left")

        btn_save_pow = tk.Button(actions_frame, text="💾 SAVE WHOLE POW", bg="#38a169", fg="white", font=("Segoe UI", 11, "bold"), padx=20, pady=5, command=self.open_project_details_modal, cursor="hand2")
        btn_save_pow.pack(side="right")

    def trigger_multi_dynamic_search(self, event):
        typed_text = self.name_entry.get().strip()
        typed_text_lower = typed_text.lower()
        
        if not typed_text:
            self.close_search_popup()
            return

        matched_results = db.search_master_items(typed_text)
        
        if not matched_results:
            self.close_search_popup()
            
            register_new = messagebox.askyesno(
                "Item Not Found", 
                f"Ang '{typed_text}' ay wala pa sa iyong database.\n\nGusto mo ba itong irehistro bilang BAGONG AYTEM?"
            )
            
            if register_new:
                self.unit_combobox.focus() # Unang ipapili ang Unit
                messagebox.showinfo("New Item Mode", "Paki-set ang Unit at Unit Price para sa bagong aytem na ito, pagkatapos ay ilagay ang Qty.")
            else:
                self.name_entry.delete(0, tk.END)
            return

        self.close_search_popup()

        # SMART PRIORITY SORTING ALGORITHM
        exact_starts = []      
        five_char_starts = []  
        somewhere_matches = [] 
        first_char_starts = [] 

        first_letter = typed_text_lower[0] if len(typed_text_lower) >= 1 else ""
        first_five = typed_text_lower[:5] if len(typed_text_lower) >= 5 else typed_text_lower

        for item in matched_results:
            name, unit, price = item[0], item[1], float(item[2])
            name_lower = name.lower()

            if name_lower.startswith(typed_text_lower):
                exact_starts.append(item)
            elif name_lower.startswith(first_five):
                five_char_starts.append(item)
            elif typed_text_lower in name_lower:
                somewhere_matches.append(item)
            elif first_letter and name_lower.startswith(first_letter):
                first_char_starts.append(item)
            else:
                somewhere_matches.append(item)

        final_sorted_results = exact_starts + five_char_starts + somewhere_matches + first_char_starts

        unique_results = []
        seen_names = set()
        for item in final_sorted_results:
            if item[0] not in seen_names:
                unique_results.append(item)
                seen_names.add(item[0])

        final_40_items = unique_results[:40] 
        total_items = len(final_40_items)
        listbox_rows = min(total_items, 40)  

        pixel_height = (listbox_rows * 22) + 4
        if pixel_height > 650: 
            pixel_height = 650

        x = self.name_entry.winfo_rootx()
        y = self.name_entry.winfo_rooty() + self.name_entry.winfo_height()
        width = self.name_entry.winfo_width() + 150

        self.search_popup = tk.Toplevel(self)
        self.search_popup.wm_overrideredirect(True)
        self.search_popup.geometry(f"{width}x{pixel_height}+{x}+{y}")
        self.search_popup.config(bg="#cbd5e0")

        self.listbox = tk.Listbox(
            self.search_popup, 
            font=("Segoe UI", 10), 
            bd=1, 
            relief="solid", 
            height=listbox_rows, 
            selectbackground="#3182ce", 
            selectforeground="white"
        )
        self.listbox.pack(fill="both", expand=True, side="left")

        popup_scroll = ttk.Scrollbar(self.search_popup, orient="vertical", command=self.listbox.yview)
        self.listbox.configure(yscrollcommand=popup_scroll.set)
        popup_scroll.pack(fill="y", side="right")

        self.master_items_cache.clear()
        for item in final_40_items:
            name, unit, price = item[0], item[1], float(item[2])
            display_text = f"{name} ({unit}) - P {price:,.2f}"
            self.listbox.insert(tk.END, display_text)
            self.master_items_cache[display_text] = {"name": name, "unit": unit, "price": price}

        if final_40_items:
            self.listbox.selection_set(0)
            self.listbox.focus_set()

        self.listbox.bind("<Return>", self.on_popup_item_select)
        self.listbox.bind("<Double-Button-1>", self.on_popup_item_select)
        self.listbox.bind("<Escape>", lambda e: self.close_search_popup())
        self.listbox.bind("<F2>", self.back_to_typing_mode)
        self.search_popup.bind("<FocusOut>", lambda e: self.after(100, self.close_search_popup))

    def on_popup_item_select(self, event=None):
        try:
            selected_index = self.listbox.curselection()[0]
            selected_text = self.listbox.get(selected_index)
            
            if selected_text in self.master_items_cache:
                item_details = self.master_items_cache[selected_text]
                
                self.name_entry.delete(0, tk.END)
                self.name_entry.insert(0, item_details['name'])
                self.unit_combobox.set(item_details['unit'])
                
                self.price_entry.delete(0, tk.END)
                self.price_entry.insert(0, str(item_details['price']))
                
                self.qty_entry.delete(0, tk.END)
                self.qty_entry.focus() # I-focus ang cursor sa Qty para mabilis mag-type
                
                self.close_search_popup()
        except IndexError:
            pass

    def close_search_popup(self):
        if self.search_popup and self.search_popup.winfo_exists():
            self.search_popup.destroy()
        self.search_popup = None
        
    def back_to_typing_mode(self, event=None):
        self.close_search_popup()
        self.name_entry.focus_set()
        self.name_entry.icursor(tk.END)

    def calculate_totals(self):
        grand_total = 0.0
        for item in self.temporary_items:
            grand_total += item['qty'] * item['price']
        self.lbl_grand_total.config(text=f"GRAND TOTAL: P {grand_total:,.2f}")

    # 🔥 REVISED & FIX: Ligtas sa ValueError at siguradong papasok sa Table Display
    def add_item_to_list(self):
        if len(self.temporary_items) >= 150:
            messagebox.showwarning("Limit Reached", "Hanggang 150 items lamang ang pwedeng ilagay sa isang POW.")
            return
        try:
            # Kunin at linisin ang input strings
            raw_qty = self.qty_entry.get().strip()
            raw_price = self.price_entry.get().strip()
            name = self.name_entry.get().strip().title()
            unit = self.unit_combobox.get().strip()

            # Pang-iwas crash: Kung walang tinype, gawing default values
            qty = int(raw_qty) if raw_qty else 0
            price = float(raw_price) if raw_price else 0.0

            if not name:
                messagebox.showwarning("Input Error", "Paki-lagay ang Item Name.")
                return

            # Check kung bago ang aytem para isulat sa Excel at DB Master list
            check_exists = db.search_master_items(name)
            is_new_item = True
            for item in check_exists:
                if item[0].lower() == name.lower():
                    is_new_item = False
                    break
            
            if is_new_item:
                if hasattr(db, 'add_new_master_item'):
                    db.add_new_master_item(name, unit, price)
                
                excel_path = r"G:\jrm\master_items.xlsx"
                try:
                    wb = load_workbook(excel_path)
                    ws = wb.active
                    new_row = ws.max_row + 1
                    ws.cell(row=new_row, column=1, value=name)
                    ws.cell(row=new_row, column=2, value=unit)
                    ws.cell(row=new_row, column=3, value=price)
                    wb.save(excel_path)
                    wb.close()
                except Exception as e:
                    print(f"[Excel Sync Warning] {e}")

            # Siguraduhing naisusulat muna ito sa active temporary array bago ang display
            self.temporary_items.append({'qty': qty, 'unit': unit, 'name': name, 'price': price})
            
            # I-update ang graphical treeview table preview
            self.refresh_table_display()
            self.calculate_totals()

            # Linisin ang forms para sa susunod na entry
            self.qty_entry.delete(0, tk.END)
            self.qty_entry.insert(0, "0")
            self.name_entry.delete(0, tk.END)
            self.price_entry.delete(0, tk.END)
            self.price_entry.insert(0, "0.00")
            
            self.name_entry.focus()
            
        except ValueError:
            messagebox.showerror("Input Error", "Dapat tamang numero ang ilagay sa Qty at Unit Price fields.")

    def remove_selected_item(self):
        selected_row = self.tree.selection()
        if not selected_row:
            messagebox.showwarning("No Selection", "Pumili muna ng aalising item sa table sa itaas.")
            return
        row_values = self.tree.item(selected_row[0], 'values')
        item_no = int(row_values[0]) - 1
        del self.temporary_items[item_no]
        self.refresh_table_display()
        self.calculate_totals()

    def refresh_table_display(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        for index, item in enumerate(self.temporary_items, start=1):
            total = item['qty'] * item['price']
            self.tree.insert("", "end", values=(index, item['qty'], item['unit'], item['name'], f"P {item['price']:,.2f}", f"P {total:,.2f}"))
        self.lbl_counter.config(text=f"Total Items: {len(self.temporary_items)} / 150")

    def open_project_details_modal(self):
        if not self.temporary_items:
            messagebox.showwarning("Empty List", "Maglagay muna ng item bago i-save.")
            return
        self.modal = tk.Toplevel(self)
        self.modal.title("Finalize POW Details")
        self.modal.geometry("400x250")
        self.modal.transient(self.winfo_toplevel())
        self.modal.grab_set()

        modal_frame = tk.Frame(self.modal, padx=20, pady=20)
        modal_frame.pack(fill="both", expand=True)

        tk.Label(modal_frame, text="Ipasok ang Pangwakas na Detalye", font=("Segoe UI", 12, "bold"), fg="#1a365d").pack(pady=(0, 15))
        tk.Label(modal_frame, text="Project Name:").pack(anchor="w")
        self.proj_name_entry = tk.Entry(modal_frame, width=50)
        self.proj_name_entry.pack(pady=5)

        tk.Label(modal_frame, text="Location:").pack(anchor="w")
        self.location_entry = tk.Entry(modal_frame, width=50)
        self.location_entry.pack(pady=5)

        btn_confirm = tk.Button(modal_frame, text="Confirm & Save to MySQL", bg="#38a169", fg="white", font=("Segoe UI", 10, "bold"), pady=5, command=self.save_to_database)
        btn_confirm.pack(fill="x", pady=15)

    def save_to_database(self):
        project_name = self.proj_name_entry.get().strip().title()
        location = self.location_entry.get().strip().title()

        if not project_name or not location:
            messagebox.showwarning("Kulang", "Ipasok ang Project Name at Location.")
            return

        success = db.save_pow_to_sql(project_name, location, self.temporary_items)
        if success:
            messagebox.showinfo("Success", "Matagumpay na nai-save ang POW!")
            self.modal.destroy()
            self.temporary_items = []
            self.refresh_table_display()
            self.lbl_grand_total.config(text="GRAND TOTAL: P 0.00")
        else:
            messagebox.showerror("Database Error", "Nagka-error sa pag-save sa MySQL.")