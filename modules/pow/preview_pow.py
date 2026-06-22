import streamlit as st
import os
import subprocess
from openpyxl import load_workbook
import db  # Koneksyon sa iyong database layer

def render_preview_pow_module():
    """
    Main component para sa pag-preview, pag-delete, at pag-edit ng Program of Works.
    100% Web-ready at compatible sa Streamlit Cloud environments.
    """
    st.markdown("## 🏛️ PREVIEW SAVED PROGRAM OF WORK (POW)")

    # ==========================================================================
    # INITIALIZATION NG SESSION STATES
    # ==========================================================================
    if 'selected_pow_id' not in st.session_state:
        st.session_state.selected_pow_id = None
    if 'in_edit_mode' not in st.session_state:
        st.session_state.in_edit_mode = False
    if 'edit_items_list' not in st.session_state:
        st.session_state.edit_items_list = []

    # Kung pinindot ang Edit, tumalon agad sa screen ng Edit Mode Interface
    if st.session_state.in_edit_mode:
        render_edit_mode_interface()
        return

    # ==========================================================================
    # MAIN SPLIT VIEW LAYOUT (Kaliwa: Project List | Kanan: Item list on POW)
    # ==========================================================================
    left_col, right_col = st.columns([1, 2])

    # --------------------------------------------------------------------------
    # 👈 KALIWANG KOLUM: List of Projects
    # --------------------------------------------------------------------------
    with left_col:
        st.markdown("#### 📋 List of Projects")
        
        try:
            projects = db.get_project_list()
        except Exception as e:
            st.error(f"Hindi makakonekta sa DB: {e}")
            projects = []

        if not projects:
            st.info("📭 Walang nahanap na proyekto sa database.")
            return

        # I-format ang listahan para maging readable dropdown options
        project_options = {f"ID {proj[0]} - {proj[1]}": proj for proj in projects}
        selected_project_key = st.radio(
            "Pumili ng Proyekto sa Listahan:",
            options=list(project_options.keys()),
            label_visibility="collapsed"
        )
        
        selected_proj_data = project_options[selected_project_key]
        pow_id = selected_proj_data[0]
        project_name = selected_proj_data[1]
        location = selected_proj_data[2]
        
        st.session_state.selected_pow_id = pow_id

        st.markdown("---")
        
        # 🗑️ PINDUTAN: Delete Entire POW
        if st.button("❌ Delete Selected POW", type="secondary", use_container_width=True):
            st.session_state.show_delete_confirmation = True
            
        if st.session_state.get('show_delete_confirmation', False):
            st.warning(f"⚠️ Sigurado ka bang buburahin ang buong proyekto: **{project_name}**? Hindi na ito mababawi.")
            conf_col1, conf_col2 = st.columns(2)
            with conf_col1:
                if st.button("👍 Oo, Burahin", type="primary", use_container_width=True):
                    success = db.delete_pow_from_sql(pow_id)
                    if success:
                        st.toast("🗑️ Matagumpay na nabura ang buong POW record.", icon="✅")
                        st.session_state.selected_pow_id = None
                        st.session_state.show_delete_confirmation = False
                        st.rerun()
            with conf_col2:
                if st.button("🔕 Kanselahin", use_container_width=True):
                    st.session_state.show_delete_confirmation = False
                    st.rerun()

    # --------------------------------------------------------------------------
    # 👉 KANAN KOLUM: Item list on POW & Summary Bar
    # --------------------------------------------------------------------------
    with right_col:
        st.markdown(f"#### 📍 Location: `{location}`")
        
        associated_items = db.get_items_by_project(pow_id)
        
        table_data = []
        grand_total = 0.0
        
        for idx, item in enumerate(associated_items, start=1):
            qty = float(item[0])
            unit = item[1]
            name = item[2]
            price = float(item[3])
            total = qty * price
            grand_total += total
            
            table_data.append({
                "#": idx,
                "Qty": qty,
                "Unit": unit,
                "Item Description": name,
                "Unit Price": f"P {price:,.2f}",
                "Total Price": f"P {total:,.2f}"
            })
            
        st.dataframe(table_data, use_container_width=True, hide_index=True)
        
        # 💰 SUMMARY BAR
        st.info(f"### **PROJECT TOTAL COST:** `P {grand_total:,.2f}`")
        
        # --- LOWER CONTAINER BUTTONS ---
        btn_col1, btn_col2 = st.columns(2)
        with btn_col1:
            if st.button("👁️ Preview Data Layout", type="primary", use_container_width=True):
                # Lumipat sa tab o module kung saan naroon ang iyong excel preview generator
                st.session_state.active_view = "Preview POW Records"
                st.toast("Inililipat ka sa Excel Preview View...", icon="👁️")
                
        with btn_col2:
            if st.button("✏️ Edit POW Record", use_container_width=True):
                st.session_state.edit_project_name = project_name
                st.session_state.edit_location = location
                # Format: [qty, unit, description, price, orig_desc]
                st.session_state.edit_items_list = [
                    [item[0], item[1], item[2], float(item[3]), item[2]] for item in associated_items
                ]
                st.session_state.in_edit_mode = True
                st.rerun()


# ==============================================================================
# ✏️ INTERNAL MODULE: THE EDIT MODE INTERFACE (Immersed Modal Emulation)
# ==============================================================================
def render_edit_mode_interface():
    st.markdown("### ✏️ EDIT MODE - Update POW Record")
    
    # --- UPPER FRAME: DETAILS ---
    with st.container(border=True):
        st.markdown("**Details of POW**")
        edit_proj_name = st.text_input("Project Title / Name:", value=st.session_state.edit_project_name)
        edit_location = st.text_input("Project Location:", value=st.session_state.edit_location)

    # --- MIDDLE FRAME: TABLE PREVIEW ---
    st.markdown("**List of Items (Current Session)**")
    display_edit_rows = []
    for idx, row in enumerate(st.session_state.edit_items_list):
        display_edit_rows.append({
            "Line": idx + 1,
            "QTY": row[0],
            "UNIT": row[1],
            "ITEM DESCRIPTION": row[2],
            "UNIT PRICE": f"{row[3]:.2f}",
            "ORIGINAL NAME (TRACER)": row[4]
        })
    st.dataframe(display_edit_rows, use_container_width=True, hide_index=True)

    # --- FORM INPUTS FOR ROWS ---
    with st.container(border=True):
        st.markdown("**Row Modification Form**")
        
        line_options = ["➕ ADD NEW LINE"] + [f"Line {i+1}: {row[2][:30]}..." for i, row in enumerate(st.session_state.edit_items_list)]
        selected_line_idx = st.selectbox("Pumili ng linyang babaguhin o magdagdag:", options=line_options)
        
        if selected_line_idx == "➕ ADD NEW LINE":
            init_qty, init_unit, init_desc, init_price, init_orig = "", "", "", "", ""
        else:
            idx = int(selected_line_idx.split(":")[0].replace("Line ", "")) - 1
            target_row = st.session_state.edit_items_list[idx]
            init_qty, init_unit, init_desc, init_price, init_orig = target_row[0], target_row[1], target_row[2], target_row[3], target_row[4]

        f_col1, f_col2, f_col3, f_col4 = st.columns([1, 1, 3, 1.5])
        with f_col1:
            form_qty = st.text_input("QTY:", value=str(init_qty), key="form_qty")
        with f_col2:
            form_unit = st.text_input("UNIT:", value=str(init_unit), key="form_unit")
        with f_col3:
            form_desc = st.text_input("DESCRIPTION:", value=str(init_desc), key="form_desc")
        with f_col4:
            form_price = st.text_input("PRICE:", value=str(init_price), key="form_price")

        action_col1, action_col2 = st.columns(2)
        with action_col1:
            if st.button("🔄 Update Selected Line", use_container_width=True, type="secondary"):
                if selected_line_idx == "➕ ADD NEW LINE":
                    st.error("Pumili muna ng valid na Line number para ma-update, boss.")
                else:
                    try:
                        idx = int(selected_line_idx.split(":")[0].replace("Line ", "")) - 1
                        st.session_state.edit_items_list[idx] = [
                            float(form_qty), form_unit, form_desc.strip(), float(form_price), init_orig
                        ]
                        st.toast("Linya matagumpay na binago!", icon="🔄")
                        st.rerun()
                    except ValueError:
                        st.error("Dapat valid na numero ang QTY at PRICE, boss.")
                        
        with action_col2:
            if st.button("➕ Add Line into Session List", use_container_width=True):
                if not form_desc.strip():
                    st.error("Lagyan ng Description ang bagong aytem, boss.")
                else:
                    try:
                        q_val = float(form_qty) if form_qty else 0.0
                        p_val = float(form_price) if form_price else 0.0
                        st.session_state.edit_items_list.append([
                            q_val, form_unit, form_desc.strip(), p_val, form_desc.strip()
                        ])
                        st.toast("Bagong linya idinagdag sa listahan!", icon="➕")
                        st.rerun()
                    except ValueError:
                        st.error("Dapat valid na numero ang QTY at PRICE, boss.")

    # --- FINAL SAVE OVERWRITE PIPELINE ---
    st.markdown("---")
    save_col1, save_col2 = st.columns(2)
    
    with save_col1:
        if st.button("💾 SAVE ALL CHANGES & OVERWRITE DATABASE", type="primary", use_container_width=True):
            excel_path = r"G:\jrm\master_items.xlsx"
            
            if not edit_proj_name.strip() or not edit_location.strip():
                st.error("Huwag iwanang blangko ang Name at Location, boss.")
                return

            # 1. EXCEL OVERWRITE ENGINE (Gagana kapag lokal ang takbo)
            if os.path.exists(excel_path):
                try:
                    wb = load_workbook(excel_path)
                    ws = wb.active
                    
                    for row in st.session_state.edit_items_list:
                        q, u, d, p, orig_d = row[0], row[1], row[2], row[3], row[4]
                        
                        target_row = None
                        for ex_row in range(2, ws.max_row + 1):
                            cell_val = ws.cell(row=ex_row, column=1).value
                            if cell_val and str(cell_val).strip().lower() == str(orig_d).strip().lower():
                                target_row = ex_row
                                break
                        
                        if target_row:
                            ws.cell(row=target_row, column=1, value=d)
                            ws.cell(row=target_row, column=2, value=u)
                            ws.cell(row=target_row, column=3, value=p)
                        else:
                            new_row = ws.max_row + 1
                            ws.cell(row=new_row, column=1, value=d)
                            ws.cell(row=new_row, column=2, value=u)
                            ws.cell(row=new_row, column=3, value=p)
                            
                    wb.save(excel_path)
                    wb.close()
                except Exception as e:
                    st.error(f"❌ Excel Sync Error: Hindi mai-save ang data sa {excel_path}. Detalye: {e}")
                    return
            else:
                # Pag nasa cloud, babasahin ito para magpatuloy ang SQL save nang walang crash
                st.warning("⚠️ Paalala: Walang nahanap na lokal na Excel drive (`G:\\`). Nilaktawan ang Excel overwrite block.")

            # 2. SUBPROCESS SCRIPT INTEGRATION
            if os.path.exists(r"G:\jrm"):
                try:
                    subprocess.run(["py", "import_master.py"], cwd=r"G:\jrm", check=True, capture_output=True, text=True)
                except Exception as err:
                    st.error(f"⚠️ Na-save sa Excel pero nagka-error ang 'import_master.py'. Detalye: {err}")
                    return

            # 3. DATABASE UPDATE REFRESH PIPELINE
            final_items_to_save = [(r[0], r[1], r[2], r[3]) for r in st.session_state.edit_items_list]
            success_main = db.update_project_main_details(st.session_state.selected_pow_id, edit_proj_name.strip(), edit_location.strip())
            success_items = db.update_project_items_batch(st.session_state.selected_pow_id, final_items_to_save)
            
            if success_main and success_items:
                st.success("🎉 Swabe ang ikot, boss! Na-save at na-overwrite na ang Database at Excel master-list.")
                st.session_state.in_edit_mode = False
                st.rerun()
            else:
                st.error("❌ SQL Error: May nagka-problema sa pag-update ng records sa database.")

    with save_col2:
        if st.button("❌ CANCEL EDIT MODE", use_container_width=True):
            st.session_state.in_edit_mode = False
            st.rerun()
